from django.shortcuts import render

# Create your views here.
from django.views import View
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, TemplateView
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q
from django.utils import timezone
from datetime import datetime, timedelta
from core.permissions import AccountantRequiredMixin
from finance.forms import FinancialTransactionForm
from .models import FinancialTransaction, IncomeCategory, ExpenseCategory
from django.core.paginator import Paginator
from datetime import datetime
from decimal import Decimal

from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.views import View
from django.db.models import Sum

from openpyxl import Workbook
from openpyxl.styles import Font

from core.permissions import AccountantRequiredMixin
from .models import CaseFinance, CaseFinanceShare, FinancialTransaction
from .forms import CaseFinanceForm, CaseFinanceShareFormSet
from cases.models import Case
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q, Sum
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from core.permissions import AccountantRequiredMixin
from .models import FinancialTransaction, IncomeCategory, ExpenseCategory
# finance/views.py
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Q

from cases.models import Case
from users.models import User              # если у тебя другая модель сотрудников — замени
from clients.models import Trustor         # если у тебя client другая модель — замени

from .models import IncomeCategory, ExpenseCategory  # замени на свои модели категорий


def _select2(queryset, label_func, q, limit=20):
    if q:
        queryset = queryset[:limit]
    else:
        queryset = queryset[:limit]
    return [{"id": obj.pk, "text": label_func(obj)} for obj in queryset]


@login_required
def ajax_cases(request):
    q = (request.GET.get("q") or "").strip()
    qs = Case.objects.all().order_by("-id")
    if q:
        qs = qs.filter(Q(title__icontains=q) | Q(case_number__icontains=q))
    results = _select2(qs, lambda o: f"#{o.id} — {o.title}", q)
    return JsonResponse({"results": results})


@login_required
def ajax_clients(request):
    q = (request.GET.get("q") or "").strip()
    qs = Trustor.objects.all().order_by("id")
    if q:
        qs = qs.filter(
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(middle_name__icontains=q) |
            Q(phone__icontains=q)
        )
    results = _select2(qs, lambda o: f"{o.get_full_name()}", q)
    return JsonResponse({"results": results})


@login_required
def ajax_employees(request):
    q = (request.GET.get("q") or "").strip()
    qs = User.objects.all().order_by("id")
    if q:
        qs = qs.filter(
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(username__icontains=q) |
            Q(email__icontains=q)
        )
    results = _select2(qs, lambda o: o.get_full_name() or o.username, q)
    return JsonResponse({"results": results})


@login_required
def ajax_income_categories(request):
    q = (request.GET.get("q") or "").strip()
    qs = IncomeCategory.objects.all().order_by("name")
    if q:
        qs = qs.filter(name__icontains=q)
    results = _select2(qs, lambda o: o.name, q)
    return JsonResponse({"results": results})


@login_required
def ajax_expense_categories(request):
    q = (request.GET.get("q") or "").strip()
    qs = ExpenseCategory.objects.all().order_by("name")
    if q:
        qs = qs.filter(name__icontains=q)
    results = _select2(qs, lambda o: o.name, q)
    return JsonResponse({"results": results})

class FinanceDashboardView(AccountantRequiredMixin, TemplateView):
    template_name = 'finance/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Обработка параметров фильтрации
        period = self.request.GET.get('period', 'month')
        transaction_type = self.request.GET.get('transaction_type', '')
        category_id = self.request.GET.get('category', '')
        date_from = self.request.GET.get('date_from', '')
        date_to = self.request.GET.get('date_to', '')
        
        # Определение периода
        today = timezone.now().date()
        
        if period == 'week':
            start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)
        elif period == 'month':
            start_date = today.replace(day=1)
            end_date = (start_date + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        elif period == 'quarter':
            current_quarter = (today.month - 1) // 3 + 1
            start_date = datetime(today.year, 3 * current_quarter - 2, 1).date()
            end_date = (datetime(today.year, 3 * current_quarter + 1, 1) - timedelta(days=1)).date()
        elif period == 'year':
            start_date = today.replace(month=1, day=1)
            end_date = today.replace(month=12, day=31)
        elif period == 'custom' and date_from and date_to:
            start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
        else:
            # По умолчанию - текущий месяц
            start_date = today.replace(day=1)
            end_date = (start_date + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        # Базовый запрос с фильтрацией по дате
        transactions = FinancialTransaction.objects.filter(date__range=[start_date, end_date])
        
        # Применяем дополнительные фильтры
        if transaction_type:
            transactions = transactions.filter(transaction_type=transaction_type)
        
        if category_id:
            # Фильтруем по категории дохода или расхода
            transactions = transactions.filter(
                Q(category_id=category_id) | Q(expense_category_id=category_id)
            )
        
        # Вычисляем общие показатели
        total_income = transactions.filter(transaction_type='income').aggregate(Sum('amount'))['amount__sum'] or 0
        total_expense = transactions.filter(transaction_type='expense').aggregate(Sum('amount'))['amount__sum'] or 0
        
        # Вычисляем рост по сравнению с предыдущим месяцем
        prev_month_start = (start_date - timedelta(days=1)).replace(day=1)
        prev_month_end = start_date - timedelta(days=1)
        
        prev_month_transactions = FinancialTransaction.objects.filter(
            date__range=[prev_month_start, prev_month_end]
        )
        prev_month_income = prev_month_transactions.filter(transaction_type='income').aggregate(Sum('amount'))['amount__sum'] or 0
        
        if prev_month_income > 0:
            growth_rate = round(((total_income - prev_month_income) / prev_month_income) * 100, 1)
        else:
            growth_rate = 0
        
        # Получаем данные для графиков
        income_by_category = transactions.filter(transaction_type='income').values(
            'category__name'
        ).annotate(total=Sum('amount')).order_by('-total')
        
        expense_by_category = transactions.filter(transaction_type='expense').values(
            'expense_category__name'
        ).annotate(total=Sum('amount')).order_by('-total')
        
        # Получаем все категории для фильтра
        income_categories = IncomeCategory.objects.all()
        expense_categories = ExpenseCategory.objects.all()
        categories = list(income_categories) + list(expense_categories)
        # Пагинация
        paginator = Paginator(transactions.order_by('-date'), 20)  # 20 элементов на страницу
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
       
        context.update({
            'page_obj': page_obj,
            'is_paginated': paginator.num_pages > 1,
            'total_income': total_income,
            'total_expense': total_expense,
            'profit': total_income - total_expense,
            'transactions': transactions.order_by('-date'),
            'income_by_category': income_by_category,
            'expense_by_category': expense_by_category,
            'categories': categories,
            'growth_rate': growth_rate,
            'start_date': start_date,
            'end_date': end_date,
            'period': period,
            'transaction_type': transaction_type,
            'category_id': category_id,
            'date_from': date_from,
            'date_to': date_to,
        })
        
        return context
    
 
 
import pandas as pd
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q
import pandas as pd
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q, Sum
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q, Sum
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from decimal import Decimal
from datetime import datetime

from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q, Sum
from django.views import View

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from core.permissions import AccountantRequiredMixin
from .models import FinancialTransaction, IncomeCategory, ExpenseCategory, CaseFinance
from cases.models import Case

from core.permissions import AccountantRequiredMixin
from .models import FinancialTransaction, IncomeCategory, ExpenseCategory
#f
class TransactionExportView(AccountantRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        qs = FinancialTransaction.objects.select_related(
            "case",
            "case__category",
            "case__finance",
            "client",
            "employee",
            "category",
            "expense_category",
            "created_by",
        )

        # --- фильтры такие же, как в списке ---
        filter_params = {}

        t_type = request.GET.get("type")
        if t_type in ("income", "expense"):
            qs = qs.filter(transaction_type=t_type)
            filter_params["Тип операции"] = "Приход" if t_type == "income" else "Расход"

        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        if date_from:
            qs = qs.filter(date__gte=date_from)
            filter_params["Дата с"] = date_from
        if date_to:
            qs = qs.filter(date__lte=date_to)
            filter_params["Дата по"] = date_to

        category_id = request.GET.get("category")
        if category_id:
            qs = qs.filter(Q(category_id=category_id) | Q(expense_category_id=category_id))
            cat = (
                IncomeCategory.objects.filter(pk=category_id).first()
                or ExpenseCategory.objects.filter(pk=category_id).first()
            )
            if cat:
                filter_params["Категория"] = cat.name

        case_id = request.GET.get("case")
        if case_id:
            qs = qs.filter(case_id=case_id)
            filter_params["Дело"] = f"ID {case_id}"

        # --- создаём книгу ---
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"financial_report_{timezone.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Финансовый отчет"

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        income_font = Font(color="00B050", bold=True)
        expense_font = Font(color="FF0000", bold=True)

        # Заголовок
        ws.merge_cells("A1:M1")
        title_cell = ws["A1"]
        title_cell.value = "Отчет по финансовым операциям"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center")

        row_num = 3
        if filter_params:
            ws.merge_cells(f"A{row_num}:M{row_num}")
            f_cell = ws[f"A{row_num}"]
            f_cell.value = "Параметры фильтрации:"
            f_cell.font = Font(bold=True)
            row_num += 1
            for k, v in filter_params.items():
                ws.merge_cells(f"A{row_num}:B{row_num}")
                cell = ws[f"A{row_num}"]
                cell.value = f"{k}: {v}"
                cell.font = Font(bold=True)
                row_num += 1
        row_num += 1  # пустая строка

        headers = [
            "Дата",               # A
            "Тип операции",       # B
            "Сумма",              # C
            "Описание",           # D
            "Категория",          # E
            "Дело",               # F
            "Доверитель (осн.)",  # G
            "№ соглашения",       # H
            "Дата соглашения",    # I
            "Клиент (поле client)",# J
            "Сотрудник",          # K
            "Создано",            # L
            "Кем создано",        # M
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_num += 1
        start_data_row = row_num

        # --- строки данных ---
        for tr in qs.order_by("-date", "-created_at"):
            case_obj = tr.case

            # Доверитель (осн.)
            main_trustor_str = "-"
            if case_obj:
                mt = case_obj.main_trustor
                if mt:
                    main_trustor_str = (
                        mt.get_full_name() if hasattr(mt, "get_full_name") else str(mt)
                    )

            # Финансы дела: номер и дата соглашения
            agreement_number = "-"
            agreement_date_str = "-"
            if case_obj:
                try:
                    finance = case_obj.finance  # OneToOne, тут может быть DoesNotExist
                except CaseFinance.DoesNotExist:
                    finance = None

                if finance:
                    if finance.agreement_number:
                        agreement_number = finance.agreement_number
                    if finance.agreement_date:
                        agreement_date_str = finance.agreement_date.strftime("%d.%m.%Y")

            # A: дата
            ws.cell(row=row_num, column=1).value = tr.date.strftime("%d.%m.%Y")

            # B: тип
            ws.cell(row=row_num, column=2).value = tr.get_transaction_type_display()

            # C: сумма
            amount_cell = ws.cell(row=row_num, column=3)
            amount_cell.value = float(tr.amount)
            amount_cell.font = income_font if tr.transaction_type == "income" else expense_font

            # D: описание
            ws.cell(row=row_num, column=4).value = tr.description

            # E: категория
            if tr.transaction_type == "income":
                cat_name = tr.category.name if tr.category else "-"
            else:
                cat_name = tr.expense_category.name if tr.expense_category else "-"
            ws.cell(row=row_num, column=5).value = cat_name

            # F: дело
            ws.cell(row=row_num, column=6).value = (
                f"#{case_obj.id} - {case_obj.title}" if case_obj else "-"
            )

            # G: доверитель (осн.)
            ws.cell(row=row_num, column=7).value = main_trustor_str

            # H: № соглашения
            ws.cell(row=row_num, column=8).value = agreement_number

            # I: дата соглашения
            ws.cell(row=row_num, column=9).value = agreement_date_str

            # J: клиент (из поля client)
            ws.cell(row=row_num, column=10).value = str(tr.client) if tr.client else "-"

            # K: сотрудник
            ws.cell(row=row_num, column=11).value = str(tr.employee) if tr.employee else "-"

            # L: создано
            ws.cell(row=row_num, column=12).value = tr.created_at.strftime("%d.%m.%Y %H:%M")

            # M: кем создано
            ws.cell(row=row_num, column=13).value = str(tr.created_by)

            # границы
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).border = border

            row_num += 1

        end_data_row = row_num - 1

        # --- итоги ---
        row_num += 1
        income_col_letter = get_column_letter(3)

        # Общий доход
        ws.merge_cells(f"A{row_num}:B{row_num}")
        c1 = ws[f"A{row_num}"]
        c1.value = "Общий доход:"
        c1.font = Font(bold=True)
        c1.fill = total_fill
        c1.border = border

        c2 = ws[f"C{row_num}"]
        c2.font = income_font
        c2.fill = total_fill
        c2.border = border
        c2.value = (
            f'=SUMIF(B{start_data_row}:B{end_data_row},"Приход",'
            f"{income_col_letter}{start_data_row}:{income_col_letter}{end_data_row})"
        )

        # Общий расход
        row_num += 1
        ws.merge_cells(f"A{row_num}:B{row_num}")
        c3 = ws[f"A{row_num}"]
        c3.value = "Общий расход:"
        c3.font = Font(bold=True)
        c3.fill = total_fill
        c3.border = border

        c4 = ws[f"C{row_num}"]
        c4.font = expense_font
        c4.fill = total_fill
        c4.border = border
        c4.value = (
            f'=SUMIF(B{start_data_row}:B{end_data_row},"Расход",'
            f"{income_col_letter}{start_data_row}:{income_col_letter}{end_data_row})"
        )

        # Чистая прибыль
        row_num += 1
        ws.merge_cells(f"A{row_num}:B{row_num}")
        c5 = ws[f"A{row_num}"]
        c5.value = "Чистая прибыль:"
        c5.font = Font(bold=True)
        c5.fill = total_fill
        c5.border = border

        c6 = ws[f"C{row_num}"]
        c6.font = Font(bold=True)
        c6.fill = total_fill
        c6.border = border
        income_row = row_num - 2
        expense_row = row_num - 1
        c6.value = f"={income_col_letter}{income_row}-{income_col_letter}{expense_row}"

        # Кол-во операций
        row_num += 2
        ws.merge_cells(f"A{row_num}:B{row_num}")
        c7 = ws[f"A{row_num}"]
        c7.value = "Количество операций:"
        c7.font = Font(bold=True)
        ws[f"C{row_num}"].value = qs.count()

        # ширина колонок
        widths = {
            "A": 12,
            "B": 14,
            "C": 12,
            "D": 40,
            "E": 20,
            "F": 28,
            "G": 22,
            "H": 16,
            "I": 16,
            "J": 20,
            "K": 20,
            "L": 18,
            "M": 20,
        }
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        # формат суммы
        for r in range(start_data_row, end_data_row + 1):
            ws.cell(row=r, column=3).number_format = '#,##0.00" сом"'

        wb.save(response)
        return response
# finance/views.py
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from cases.models import Case

@login_required
def ajax_case_stages(request):
    case_id = request.GET.get('case_id')
    if not case_id:
        return JsonResponse({'stages': []})

    case = Case.objects.select_related('category').filter(pk=case_id).first()
    if not case or not case.category_id:
        return JsonResponse({'stages': []})

    stages = [
        {'id': s.id, 'name': s.name}
        for s in case.category.stages.all().order_by('order')
    ]
    return JsonResponse({'stages': stages})

class TransactionListView(AccountantRequiredMixin, ListView):
    model = FinancialTransaction
    template_name = 'finance/transaction_list.html'
    context_object_name = 'transactions'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Фильтрация по типу операции
        transaction_type = self.request.GET.get('type')
        if transaction_type in ['income', 'expense']:
            queryset = queryset.filter(transaction_type=transaction_type)
        
        # Фильтрация по дате
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        
        # Фильтрация по категории
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(Q(category_id=category) | Q(expense_category_id=category))
        
        # Фильтрация по делу
        case = self.request.GET.get('case')
        if case:
            queryset = queryset.filter(case_id=case)
        
        return queryset.order_by('-date', '-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Получаем отфильтрованный queryset
        queryset = self.get_queryset()
        
        # Вычисляем общую статистику
        total_income = queryset.filter(transaction_type='income').aggregate(Sum('amount'))['amount__sum'] or 0
        total_expense = queryset.filter(transaction_type='expense').aggregate(Sum('amount'))['amount__sum'] or 0
        net_income = total_income - total_expense
        
        context.update({
            'income_categories': IncomeCategory.objects.all(),
            'expense_categories': ExpenseCategory.objects.all(),
            'total_income': total_income,
            'total_expense': total_expense,
            'net_income': net_income,
        })
        
        return context   


class TransactionCreateView(AccountantRequiredMixin, CreateView):
    model = FinancialTransaction
    form_class = FinancialTransactionForm
    template_name = 'finance/transaction_form.html'
    success_url = reverse_lazy('finance:transaction_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)

class TransactionUpdateView(AccountantRequiredMixin, UpdateView):
    model = FinancialTransaction
    template_name = 'finance/transaction_form.html'
    fields = [
        'transaction_type', 'amount', 'date', 'description',
        'category', 'expense_category', 'case', 'stage', 'client', 'employee'
    ]
    
    def get_success_url(self):
        return reverse_lazy('finance:transaction_list')

class TransactionDeleteView(AccountantRequiredMixin, DeleteView):
    model = FinancialTransaction
    template_name = 'finance/transaction_confirm_delete.html'
    success_url = reverse_lazy('finance:transaction_list')
    
    
# finance/views.py (добавь импортов)
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse
from django.db import transaction
from django.contrib import messages

from .models import CaseFinance
from .forms import CaseFinanceForm, CaseFinanceShareFormSet
from core.permissions import AccountantRequiredMixin  # как у тебя в Dashboard


# class CaseFinanceUpdateView(AccountantRequiredMixin, View):
#     """
#     Редактирование финансов по делу + распределение 70% между сотрудниками.
#     URL вида: /finance/cases/<int:case_id>/
#     """
#     template_name = 'finance/case_finance_form.html'

#     def get_object(self, case_id):
#         finance = get_object_or_404(CaseFinance, case_id=case_id)
#         return finance

#     @method_decorator(login_required)
#     def dispatch(self, request, *args, **kwargs):
#         return super().dispatch(request, *args, **kwargs)

#     def get(self, request, case_id):
#         finance = self.get_object(case_id)
#         form = CaseFinanceForm(instance=finance)
#         formset = CaseFinanceShareFormSet(instance=finance)

#         return render(request, self.template_name, {
#             'finance': finance,
#             'form': form,
#             'formset': formset,
#         })

#     def post(self, request, case_id):
#         finance = self.get_object(case_id)
#         form = CaseFinanceForm(request.POST, instance=finance)
#         formset = CaseFinanceShareFormSet(request.POST, instance=finance)

#         if form.is_valid() and formset.is_valid():
#             with transaction.atomic():
#                 finance = form.save()

#                 # Сохраняем доли
#                 shares = formset.save(commit=False)

#                 # удаление отмеченных
#                 for obj in formset.deleted_objects:
#                     obj.delete()

#                 for share in shares:
#                     share.case_finance = finance
#                     share.save()

#                 # Пересчитываем суммы
#                 finance.recalc_shares()

#             messages.success(request, 'Финансовая карточка дела успешно обновлена.')
#             # редирект, например, обратно в дело
#             return redirect(reverse('cases:case_detail', args=[finance.case_id]))

#         return render(request, self.template_name, {
#             'finance': finance,
#             'form': form,
#             'formset': formset,
#         })

# class CaseFinanceUpdateView(AccountantRequiredMixin, View):
#     template_name = 'finance/case_finance_form.html'

#     @method_decorator(login_required)
#     def get(self, request, case_id):
#         case = get_object_or_404(Case, pk=case_id)
#         from decimal import Decimal
#         finance, created = CaseFinance.objects.get_or_create(
#             case=case,
#             defaults={
#                 'contract_amount': case.contract_amount or Decimal('0.00'),
#                 'paid_amount': Decimal('0.00'),
#             }
#         )

#         # Если карточка есть, но нет долей — создаём по ответственным юристам поровну
#         if not finance.shares.exists():
#             lawyers = case.responsible_lawyer.all()
#             count = lawyers.count()
#             if count > 0:
#                 base = (Decimal('100.00') / count).quantize(Decimal('0.01'))
#                 remainder = Decimal('100.00') - base * count
#                 for idx, lawyer in enumerate(lawyers):
#                     percent = base + remainder if idx == 0 else base
#                     CaseFinanceShare.objects.create(
#                         case_finance=finance,
#                         employee=lawyer,
#                         percent_of_pool=percent,
#                     )
#                 finance.recalc_shares()

#         form = CaseFinanceForm(instance=finance)
#         formset = CaseFinanceShareFormSet(instance=finance)

#         return render(request, self.template_name, {
#             'case': case,
#             'finance': finance,
#             'form': form,
#             'formset': formset,
#         })

#     @method_decorator(login_required)
#     def post(self, request, case_id):
#         case = get_object_or_404(Case, pk=case_id)
#         from decimal import Decimal
#         finance, created = CaseFinance.objects.get_or_create(
#             case=case,
#             defaults={
#                 'contract_amount': case.contract_amount or Decimal('0.00'),
#                 'paid_amount': Decimal('0.00'),
#             }
#         )

#         form = CaseFinanceForm(request.POST, instance=finance)
#         formset = CaseFinanceShareFormSet(request.POST, instance=finance)

#         if form.is_valid() and formset.is_valid():
#             finance = form.save()
#             formset.save()
#             finance.recalc_shares()
#             messages.success(request, 'Финансы по делу успешно обновлены.')
#             return redirect('cases:case_detail', pk=case.pk)

#         return render(request, self.template_name, {
#             'case': case,
#             'finance': finance,
#             'form': form,
#             'formset': formset,
#         })
# finance/views.py

from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.shortcuts import get_object_or_404, render, redirect
from django.views import View

from core.permissions import AccountantRequiredMixin
from cases.models import Case
from .models import (
    FinancialTransaction, IncomeCategory, ExpenseCategory,
    CaseFinance, CaseFinanceShare
)
from .forms import (
    FinancialTransactionForm,
    CaseFinanceForm,
    CaseFinanceShareFormSet,
)


class CaseFinanceUpdateView(AccountantRequiredMixin, View):
    """
    Редактирование финансовой карточки дела:
    - Номер/дата соглашения
    - Сумма договора, оплачено
    - Проценты 30% / 70%
    - Доли сотрудников в 70%
    """
    template_name = "finance/case_finance_form.html"

    def _get_case_and_finance(self, case_id):
        case = get_object_or_404(Case, pk=case_id)

        finance, created = CaseFinance.objects.get_or_create(
            case=case,
            defaults={
                "contract_amount": Decimal("0.00"),
                "paid_amount": Decimal("0.00"),
            },
        )

        # Если только что создали фин. карточку — раскидываем 70% поровну
        if created:
            lawyers_qs = case.responsible_lawyer.all()
            count = lawyers_qs.count()
            if count > 0:
                base = (Decimal("100.00") / count).quantize(Decimal("0.01"))
                remainder = Decimal("100.00") - base * count

                shares = []
                for idx, lawyer in enumerate(lawyers_qs):
                    percent = base + remainder if idx == 0 else base
                    shares.append(
                        CaseFinanceShare(
                            case_finance=finance,
                            employee=lawyer,
                            percent_of_pool=percent,
                        )
                    )
                CaseFinanceShare.objects.bulk_create(shares)
                finance.recalc_shares()

        return case, finance

    @method_decorator(login_required)
    def get(self, request, case_id):
        case, finance = self._get_case_and_finance(case_id)
        form = CaseFinanceForm(instance=finance)
        formset = CaseFinanceShareFormSet(instance=finance)

        context = {
            "case": case,
            "finance": finance,
            "form": form,
            "formset": formset,
        }
        return render(request, self.template_name, context)

    @method_decorator(login_required)
    def post(self, request, case_id):
        case, finance = self._get_case_and_finance(case_id)
        form = CaseFinanceForm(request.POST, instance=finance)
        formset = CaseFinanceShareFormSet(request.POST, instance=finance)

        if form.is_valid() and formset.is_valid():
            finance = form.save()
            formset.save()
            finance.recalc_shares()
            messages.success(request, "Финансовая карточка дела сохранена.")
            return redirect("finance:case_finance", case_id=case.id)

        # Если есть ошибки — просто показываем их на этой же странице
        context = {
            "case": case,
            "finance": finance,
            "form": form,
            "formset": formset,
        }
        return render(request, self.template_name, context)

# class CaseFinanceUpdateView(AccountantRequiredMixin, View):
#     template_name = 'finance/case_finance_form.html'

#     @method_decorator(login_required)
#     def get(self, request, case_id):
#         from decimal import Decimal
#         case = get_object_or_404(Case, pk=case_id)
#         finance, created = CaseFinance.objects.get_or_create(
#             case=case,
#             defaults={
#                 'contract_amount': case.contract_amount or Decimal('0.00'),
#                 'paid_amount': Decimal('0.00'),
#             }
#         )

#         # если нет долей – создаём по ответственным юристам поровну
#         if not finance.shares.exists():
#             lawyers = case.responsible_lawyer.all()
#             count = lawyers.count()
#             if count > 0:
#                 base = (Decimal('100.00') / count).quantize(Decimal('0.01'))
#                 remainder = Decimal('100.00') - base * count
#                 for idx, lawyer in enumerate(lawyers):
#                     percent = base + remainder if idx == 0 else base
#                     CaseFinanceShare.objects.create(
#                         case_finance=finance,
#                         employee=lawyer,
#                         percent_of_pool=percent,
#                     )
#                 finance.recalc_shares()

#         form = CaseFinanceForm(instance=finance)
#         formset = CaseFinanceShareFormSet(instance=finance)

#         return render(request, self.template_name, {
#             'case': case,
#             'finance': finance,
#             'form': form,
#             'formset': formset,
#         })

#     @method_decorator(login_required)
#     def post(self, request, case_id):
#         from decimal import Decimal
#         case = get_object_or_404(Case, pk=case_id)
#         finance, created = CaseFinance.objects.get_or_create(
#             case=case,
#             defaults={
#                 'contract_amount': case.contract_amount or Decimal('0.00'),
#                 'paid_amount': Decimal('0.00'),
#             }
#         )

#         form = CaseFinanceForm(request.POST, instance=finance)
#         formset = CaseFinanceShareFormSet(request.POST, instance=finance)

#         if form.is_valid() and formset.is_valid():
#             finance = form.save()
#             formset.save()
#             finance.recalc_shares()

#             # Синхронизируем сумму договора в самом деле
#             case.contract_amount = finance.contract_amount
#             case.save(update_fields=['contract_amount', 'updated_at'])

#             messages.success(request, 'Финансы по делу успешно обновлены.')
#             return redirect('cases:case_detail', pk=case.pk)

#         return render(request, self.template_name, {
#             'case': case,
#             'finance': finance,
#             'form': form,
#             'formset': formset,
#         })

# finance/views.py (ниже TransactionExportView)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime
from .models import CaseFinance
from decimal import Decimal
from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.db.models import Sum
from django.http import HttpResponse
from django.views import View

from openpyxl import Workbook
from openpyxl.styles import Font

from core.permissions import AccountantRequiredMixin
from .models import CaseFinance, FinancialTransaction


class CaseFinanceExportView(AccountantRequiredMixin, View):
    """
    Экспорт: одна строка = одно дело.
    Все операции по делу (приходы) суммируются в одно поле.
    """

    @method_decorator(login_required)
    def get(self, request):
        qs = CaseFinance.objects.select_related("case")

        # Фильтр по датам создания дела (если нужно)
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        if date_from:
            qs = qs.filter(case__created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(case__created_at__date__lte=date_to)

        wb = Workbook()
        ws = wb.active
        ws.title = "Финансы по делам"

        headers = [
            "ID дела",
            "Дата создания дела",
            "Название дела",
            "Доверитель (основной)",
            "Номер соглашения",
            "Дата соглашения",
            "Сумма договора (100%)",
            "Оплачено по договору",
            "Сумма операций (приходы)",
            "30% компании",
            "70% юристов",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for finance in qs:
            case = finance.case

            # Основной доверитель — через property case.main_trustor
            trustor = case.main_trustor
            if trustor:
                # у Trustor есть get_full_name() и __str__
                trustor_name = (
                    trustor.get_full_name()
                    if hasattr(trustor, "get_full_name")
                    else str(trustor)
                )
            else:
                trustor_name = ""

            # ВСЕ приходы по делу (могут отличаться от paid_amount, если что-то правили руками)
            total_income = (
                FinancialTransaction.objects.filter(
                    case=case,
                    transaction_type="income",
                ).aggregate(total=Sum("amount"))["total"]
                or Decimal("0.00")
            )

            ws.append(
                [
                    case.id,
                    case.created_at.strftime("%d.%m.%Y") if case.created_at else "",
                    case.title,
                    trustor_name,
                    finance.agreement_number or "",
                    finance.agreement_date.strftime("%d.%m.%Y")
                    if finance.agreement_date
                    else "",
                    float(finance.contract_amount or 0),
                    float(finance.paid_amount or 0),
                    float(total_income),
                    float(finance.company_share_amount),
                    float(finance.lawyers_pool_amount),
                ]
            )

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        )
        filename = f"case_finance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

# class CaseFinanceExportView(AccountantRequiredMixin, View):
#     """
#     Экспорт: одна строка = одно дело.
#     Все операции по делу (приходы) суммируются в одно поле.
#     """
#     @method_decorator(login_required)
#     def get(self, request):
#         qs = CaseFinance.objects.select_related('case')

#         # Фильтр по датам, если надо
#         date_from = request.GET.get('date_from')
#         date_to = request.GET.get('date_to')
#         if date_from:
#             qs = qs.filter(case__created_at__date__gte=date_from)
#         if date_to:
#             qs = qs.filter(case__created_at__date__lte=date_to)

#         wb = Workbook()
#         ws = wb.active
#         ws.title = 'Финансы по делам'

#         headers = [
#             'ID дела',
#             'Дата создания дела',
#             'Название дела',
#             'Доверитель (основной)',
#             'Номер соглашения',
#             'Дата соглашения',
#             'Сумма договора (100%)',
#             'Оплачено по договору',
#             'Сумма операций (приходы)',  # вот тут ВСЕ операции по делу
#             '30% компании',
#             '70% юристов',
#         ]
#         ws.append(headers)
#         for cell in ws[1]:
#             cell.font = Font(bold=True)

#         for finance in qs:
#             case = finance.case
#             # основной доверитель
#             try:
#                 trustor = case.main_trustor
#             except Exception:
#                 trustor = None
#             trustor_name = getattr(trustor, 'full_name', '') if trustor else ''

#             total_income = FinancialTransaction.objects.filter(
#                 case=case,
#                 transaction_type='income',
#             ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

#             ws.append([
#                 case.id,
#                 case.created_at.strftime('%d.%m.%Y') if case.created_at else '',
#                 case.title,
#                 trustor_name,
#                 finance.agreement_number,
#                 finance.agreement_date.strftime('%d.%m.%Y') if finance.agreement_date else '',
#                 float(finance.contract_amount),
#                 float(finance.paid_amount),
#                 float(total_income),  # ВСЕ ПРИХОДЫ ПО ДЕЛУ
#                 float(finance.company_share_amount),
#                 float(finance.lawyers_pool_amount),
#             ])

#         response = HttpResponse(
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         filename = f'case_finance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
#         response['Content-Disposition'] = f'attachment; filename="{filename}"'
#         wb.save(response)
#         return response
# class CaseFinanceExportView(AccountantRequiredMixin, View):
#     """
#     Экспорт дел с финансовыми данными в Excel.
#     Фильтр по дате создания дела: ?date_from=YYYY-MM-DD&date_to=YYYY-MM-DD
#     """
#     def get(self, request, *args, **kwargs):
#         qs = CaseFinance.objects.select_related('case').prefetch_related('shares__employee')

#         date_from = request.GET.get('date_from')
#         date_to = request.GET.get('date_to')

#         if date_from:
#             qs = qs.filter(case__created_at__date__gte=date_from)
#         if date_to:
#             qs = qs.filter(case__created_at__date__lte=date_to)

#         # создаём Excel
#         wb = Workbook()
#         ws = wb.active
#         ws.title = 'Дела'

#         bold = Font(bold=True)
#         border = Border(
#             left=Side(style='thin'),
#             right=Side(style='thin'),
#             top=Side(style='thin'),
#             bottom=Side(style='thin')
#         )

#         headers = [
#             'ID дела',
#             'Номер дела',
#             'Дата создания дела',
#             'Номер соглашения',
#             'Дата соглашения',
#             'Доверители',
#             'Сумма договора',
#             'Оплачено',
#             '30% компании',
#             '70% юристов',
#             'Распределение 70% (ФИО: % / сумма полностью / сумма сейчас)',
#         ]

#         row = 1
#         for col, h in enumerate(headers, start=1):
#             cell = ws.cell(row=row, column=col, value=h)
#             cell.font = bold
#             cell.border = border

#         row += 1

#         for cf in qs:
#             case = cf.case

#             trustors = ", ".join([str(t) for t in case.all_trustors])  # твой property
#             shares_strings = []
#             for share in cf.shares.all():
#                 s = f"{share.employee.get_full_name() or share.employee.username}: {share.percent_of_pool}% / {share.amount_full} / {share.amount_current}"
#                 shares_strings.append(s)
#             shares_text = "; ".join(shares_strings) if shares_strings else "-"

#             data = [
#                 case.id,
#                 case.case_number or '',
#                 case.created_at.strftime('%d.%m.%Y %H:%M'),
#                 cf.agreement_number or '',
#                 cf.agreement_date.strftime('%d.%m.%Y') if cf.agreement_date else '',
#                 trustors,
#                 float(cf.contract_amount or 0),
#                 float(cf.paid_amount or 0),
#                 float(cf.company_share_amount),
#                 float(cf.lawyers_pool_amount),
#                 shares_text,
#             ]

#             for col, value in enumerate(data, start=1):
#                 cell = ws.cell(row=row, column=col, value=value)
#                 cell.border = border

#             row += 1

#         response = HttpResponse(
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         filename = f"cases_finance_{timezone.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
#         response['Content-Disposition'] = f'attachment; filename="{filename}"'

#         wb.save(response)
#         return response
